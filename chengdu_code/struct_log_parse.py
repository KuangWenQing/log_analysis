#!/usr/bin/env python3
import sys
import os
import openpyxl


def write_excel_xlsx(sheet, value, row_cnt=0):
    """ sheet 工作簿
        value 是写的内容
        row_cnt 表示写到表格第几行"""
    index = len(value)
    for i in range(0, index):
        row_cnt += 1
        for j in range(0, len(value[i])):
            sheet.cell(row=row_cnt, column=j+1, value=str(value[i][j]))
    return row_cnt


num_argv = len(sys.argv)
if num_argv < 2:
    path = "/home/kwq/chengdu/data/bin/"
    file = "0310"
else:
    path = sys.argv[1]
    file = sys.argv[2]

path_file = path + file
'''
gga = "$GPGGA,024028.00,4000.0013,N,11559.9974,E,1,10,1.22,113.714,M,-8.00,M,,*78"
'''
Lamp_dict = {1: "D1", 2: "D11", 3: "D21", 4: "D31", 5: "D51", 6: "D61", 7: "D71", 8: "D81", 9: "D91", 10: "D101"}


def excel_row_data(data_lst, Id):
    avepli, avecnr, ircnt, nval, warn_pv_rate_0, warn_pv_rate_1, avensv = 0, 0, 0, 0, 0, 0, 0
    for info in data_lst:
        if "fall" in info:
            if "avepli" in info:
                avepli = int(info.split('=')[-1])
            elif "avecnr" in info:
                avecnr = int(info.split('=')[-1])
            elif "ircnt" in info:
                ircnt = int(info.split('=')[-1])
            elif "nval" in info:
                nval = int(info.split('=')[-1])
            elif "warn_pv_rate_0" in info:
                warn_pv_rate_0 = float(info.split('=')[-1])
            elif "warn_pv_rate_1" in info:
                warn_pv_rate_1 = float(info.split('=')[-1])
            elif "avensv" in info:
                avensv = float(info.split('=')[-1])
    write_row_data = [Id, avepli, avecnr, ircnt, nval, warn_pv_rate_0, warn_pv_rate_1, avensv]
    return write_row_data


def excel_write(write_value):
    head_xlsx = [["灯杆号", "ave_pli", "ave_cnr", "ir_cnt", "pos_valid", "warn_pv_rate_0", "warn_pv_rate_1", "ave_sv",
                  "picture", "kml"], [file]]
    book_name_xlsx = path + 'summary_table.xlsx'
    try:
        wb = openpyxl.load_workbook(book_name_xlsx)
        ws = wb.active
        row = 2
        while ws.cell(row, 1).value:
            row += 1
        row_xlsx = write_excel_xlsx(ws, [[file]], row-1)
    except:
        wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
        ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
        row_xlsx = write_excel_xlsx(ws, head_xlsx, row_cnt=0)

    row_renew = write_excel_xlsx(ws, write_value, row_xlsx)
    wb.save(book_name_xlsx)
    wb.close()
    return row_renew


def gga_produce(data_lst, gga_time="010038.00,"):
    gga_head = "$GPGGA,"
    gga_tail = ",E,1,10,1.22,113.714,M,-8.00,M,,*78"
    hour = int(gga_time[:2]) + 1
    if hour > 23:
        hour = 0
    if hour < 10:
        gga_time = '0' + str(hour) + gga_time[2:]
    else:
        gga_time = str(hour) + gga_time[2:]
    lon_str = data_lst[1].split('=')[-1]
    lat_str = data_lst[2].split('=')[-1]

    tmp_idx = lon_str.index('.')
    minute = round(float(lon_str[tmp_idx:]) * 60, 4)
    if minute < 10:
        lon_gga = lon_str[:tmp_idx].lstrip() + '0' + str(minute)
    else:
        lon_gga = lon_str[:tmp_idx].lstrip() + str(minute)

    tmp_idx = lat_str.index('.')
    minute = round(float(lat_str[tmp_idx:]) * 60, 4)
    if minute < 10:
        lat_gga = lat_str[:tmp_idx].lstrip() + '0' + str(minute)
    else:
        lat_gga = lat_str[:tmp_idx].lstrip() + str(minute)

    GGA = gga_head + gga_time + lat_gga + ',N,' + lon_gga + gga_tail
    return GGA, gga_time


def run():
    gga_time = "010038.00,"
    gga_file = open(path_file+".gga", 'w')
    write_data = []

    with open(path_file, 'r') as fd:
        row_cnt = 0
        for row in fd:
            row_cnt += 1
            if len(row) < 11:
                continue
            s = ''
            try:
                s = row.split('47 50 53 3A ')[1]
            except:
                print("bin data is abnormal")
                continue

            if s:
                exe = '/home/kwq/project/py/log_analysis/chengdu_code/struct_parse ' + s
                p = os.popen(exe)
                data = p.read()
                data_list = data.split('\n')
                gga, gga_time = gga_produce(data_list, gga_time)
                print(gga, file=gga_file)
                write_data.append(excel_row_data(data_list, Lamp_dict[row_cnt]))

    excel_write(write_data)


if __name__ == "__main__":
    run()
