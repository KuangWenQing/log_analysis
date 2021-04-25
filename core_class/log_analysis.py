import openpyxl
import numpy as np
from log_parser import LogParser
import matplotlib.pyplot as plt
from base_function import ecef_to_enu, write_excel_xlsx, find_abnormal_data, xyz_to_lla, degree_to_dms

__all__ = ['LogAnalysis']


class LogAnalysis(LogParser):
    def __init__(self, target_file: str, __purpose__: list, ubx_file=''):
        super().__init__(target_file, __purpose__, ubx_file)

    def final_pos_analysis(self, true_xyz, fd_st):
        original_str = self.final_pos_str
        print(original_str)
        try:
            str_xyz = original_str[original_str.index('xyz,') + 5: original_str.index(', idx')]
        except:
            if fd_st:
                print("no | DEBUG R AVE", end='|', file=fd_st)
            return 0
        str_xyz = str_xyz.split(',')
        final_xyz = list([float(x) for x in str_xyz])
        final_lla = xyz_to_lla(final_xyz[0], final_xyz[1], final_xyz[2])
        final_enu = ecef_to_enu(true_xyz[0], true_xyz[1], true_xyz[2], final_lla[0], final_lla[1], final_lla[2])
        print("\n8088 最终定位位置:", final_lla)
        final_lat = degree_to_dms(str(final_lla[0]))[0]
        final_lon = degree_to_dms(str(final_lla[1]))[0]
        print(final_lat, final_lon, final_lla[2])
        print("东北天 = ", final_enu)

        dis_en = np.linalg.norm(np.array([final_enu[0], final_enu[1]]))
        dis_T_xyz = np.linalg.norm(np.array(final_xyz) - np.array(true_xyz))
        print("平面误差 dis_en = ", round(dis_en, 3))
        print("与真值的距离 dis_T_xyz = ", round(dis_T_xyz, 3))
        if fd_st:
            print("{:.3f}|{:.3f}".format(dis_T_xyz, dis_en), end='|', file=fd_st)

    def pli_abnormal_pli_mean_cnr_mean(self, fd_st):
        _fd_ = open(self.path + 'chart/' + self.filename.split('.')[0] + '_ab_pli.txt', 'w')
        pli_mean = []
        cnr_mean = []
        time_lst = []
        pli_row_bak = []
        sum_bak = 100
        abnormal_pli_cnt = 0
        for per_sec_info in self.all_info_list:
            pli = []
            idx = []
            try:
                _tmp_pli_lst = per_sec_info['pli']
            except:
                continue
            # for i, item in enumerate(_tmp_pli_lst):
            #     if item != 100:
            #         pli.append(item)
            #         idx.append(i)
            pli, idx = self.valid_pli_and_index(_tmp_pli_lst)
            if len(pli):
                pli_mean.append(np.mean(np.array(pli)))
                pli_sum = np.sum(pli)
                if pli_sum / sum_bak > 2 and pli_mean[-1] > 15 and len(pli) > 2:
                    print(pli_row_bak, per_sec_info['chl_time'], per_sec_info['pli'], file=_fd_)
                    abnormal_pli_cnt += 1
                sum_bak = pli_sum
            else:
                sum_bak = 100
                pli_mean.append(0)
            pli_row_bak = per_sec_info['pli']

            if idx:
                cnr = []
                for i in idx:
                    cnr.append(per_sec_info['cnr'][i])
                cnr_mean.append(np.mean(np.array(cnr)))
            else:
                cnr_mean.append(60)

            time_lst.append(per_sec_info['chl_time'])

        _fd_.close()
        print("abnormal pli rate = {:.2%}".format(abnormal_pli_cnt / len(pli_mean)))
        print("mean(pli_mean) = {:.2f}, std(pli_mean) = {:.2f}".format(np.mean(pli_mean), np.std(pli_mean)))
        print("mean(cnr_mean) = {:.2f}, std(cnr_mean) = {:.2f}".format(np.mean(cnr_mean), np.std(cnr_mean)))
        if fd_st:
            print("{:.3f}|{:.3f}|{:.3%}|{:.3f}|{:.3f}".format(
                np.mean(pli_mean), np.std(pli_mean), abnormal_pli_cnt / len(pli_mean),
                np.mean(cnr_mean), np.std(cnr_mean)), end='|', file=fd_st)

        fig1 = plt.figure(1)
        plt.title("per sec pli list's mean and cnr list's mean")
        plt.plot(time_lst, pli_mean, marker='*', label='pli')
        plt.plot(time_lst, cnr_mean, marker='o', label='cnr')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + "chart/" + self.filename[:-4] + "_per_sec_pli_cnr.png")
        # plt.show()
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

    def static_pos_cmp(self, true_xyz, fd_st=None):
        RMC_list = self.all_info_dict['RMC']
        time_list = self.all_info_dict['chl_time']
        all_diff_xyz = []
        all_EN = []
        time_pos = []
        diff_vel = []
        time_vel = []
        gga_1_rmc_N = 0
        sec = 0
        for info in self.all_info_dict['GGA']:
            rmc_dict = RMC_list[sec]
            SoW = time_list[sec]  # second of week

            if info['valid']:
                diff_xyz = (np.array(true_xyz) - np.array(info['xyz']))
                all_diff_xyz.append(diff_xyz)  # 收集点(x,y,z)3个轴的误差
                ENU = ecef_to_enu(true_xyz[0], true_xyz[1], true_xyz[2], info['lat'], info['lon'], info['high'])
                all_EN.append(ENU[:2])  # 收集东北
                time_pos.append(SoW)

            if rmc_dict['valid']:
                diff_vel.append(rmc_dict['speed'])
                time_vel.append(SoW)
            else:
                if info['valid']:
                    gga_1_rmc_N += 1
            sec += 1

        all_dis_xyz = [np.linalg.norm(diff_xyz) for diff_xyz in all_diff_xyz]
        all_dis_EN = [np.linalg.norm(en) for en in all_EN]
        len_tmp = len(time_list)
        print("fix rate = {:.3%}, warning rate = {:.3%}".format(len(time_pos) / len_tmp, gga_1_rmc_N / len_tmp))
        self.sort_and_print_50_95_99(all_dis_xyz, "all_dis_xyz", fd_st)
        max_dis = self.sort_and_print_50_95_99(all_dis_EN, "all_dis_EN", fd_st)
        self.sort_and_print_50_95_99(diff_vel, "diff_vel", fd_st)
        if fd_st:
            print("{:.3%}|{:.3%}".format(len(time_pos) / len_tmp, gga_1_rmc_N / len_tmp),
                  end='|', file=fd_st)

        fig1 = plt.figure(1)
        plt.title("distance with true position")
        plt.plot(time_pos, all_dis_xyz, marker='*', label='dis_xyz')
        plt.plot(time_pos, all_dis_EN, marker='o', label='dis_EN')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_pos.png')
        if max_dis > 100:
            plt.show()
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

    def _cmp_with_true_pos_(self, true_xyz, aim="GGA"):
        time_list = self.all_info_dict['chl_time']
        all_diff_xyz = []
        all_EN = []
        time_lst = []
        sec = 0
        for info in self.all_info_dict[aim]:

            SoW = time_list[sec]  # second of week

            if info['valid']:
                diff_xyz = (np.array(true_xyz) - np.array(info['xyz']))
                all_diff_xyz.append(diff_xyz)  # 收集点(x,y,z)3个轴的误差
                ENU = ecef_to_enu(true_xyz[0], true_xyz[1], true_xyz[2], info['lat'], info['lon'], info['high'])
                all_EN.append(ENU[:2])  # 收集东北
                time_lst.append(SoW)
            sec += 1

        all_dis_xyz = [np.linalg.norm(diff_xyz) for diff_xyz in all_diff_xyz]
        all_dis_EN = [np.linalg.norm(en) for en in all_EN]
        return time_lst, all_dis_EN, all_dis_xyz

    def cmp_with_true_draw_picture(self, true_xyz, target=["GGA", ]):
        """
        :param true_xyz:
        :param target: list of GGA, such as : [GPGGA, BDGGA, GPGGA KF]
        """
        dict_tmp = {}
        for aim in target:
            dict_tmp[aim] = dict(zip(["time", "EN", "ENU"], self._cmp_with_true_pos_(true_xyz, aim)))

        str_label = '_'
        for key in dict_tmp.keys():
            self.sort_and_print_50_95_99(dict_tmp[key]["ENU"], key + "_ENU")
            self.sort_and_print_50_95_99(dict_tmp[key]["EN"], key + "_EN")
            str_label += key + "_"
        self.draw_in_one_pic(dict_tmp, save_name=str_label)

    def draw_in_one_pic(self, target_dict, save_name=''):
        fig1 = plt.figure(1)
        plt.suptitle("distance with true position")
        plt.subplot(211)
        plt.title('ENU')
        for key in target_dict.keys():
            plt.plot(target_dict[key]["time"], target_dict[key]["ENU"], marker='*', label=key)
        plt.legend()  # 不加该语句无法显示 label

        plt.subplot(212)
        plt.title('EN')
        for key in target_dict.keys():
            plt.plot(target_dict[key]["time"], target_dict[key]["EN"], marker='o', label=key)
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + save_name + '_cmp_pos.png')
        plt.pause(10)  # 间隔的秒数： 11s
        plt.close(fig1)

    def sort_and_print_50_95_99(self, aim_list, keyword, fd_st=None):
        sort_list = np.sort(aim_list)
        len_tmp = len(sort_list)
        percentage_50 = sort_list[int(len_tmp * 0.5)]
        percentage_95 = sort_list[int(len_tmp * 0.95)]
        percentage_99 = sort_list[int(len_tmp * 0.99)]
        std_list = np.std(sort_list)
        print(keyword + " 50% = {:f}, 95% = {:f}, 99% = {:f}, std = {:f}"
              .format(percentage_50, percentage_95, percentage_99, std_list))
        if fd_st:
            print("{:.3f}|{:.3f}|{:.3f}|{:.3f}".format(percentage_50, percentage_95, percentage_99, std_list),
                  end='|', file=fd_st)
        return sort_list[-1]

    # noinspection PyTypeChecker
    def cnr_cmp(self, fd_st):
        if not self.cmp_support or not self.cmp_enable:
            return
        head_xlsx = [['说明', "[F9P每一秒的CNR]该列表的均值和标准差", '', "[8088每一秒的CNR]该列表的均值和标准差", '',
                      "[每一秒各个卫星的载噪比与F9P的载噪比之差] 该列表 的均值和标准差"],
                     ['', 'f9p_ave', 'f9p_std', '8088_ave', '8088_std', 'f9p-8088(ave)', 'f9p-8088(std)', 'cnt'],
                     [self.filename, ]]
        book_name_xlsx = self.path + 'chart/' + '_compare_cnr.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 3
            while ws.cell(row, 1).value:
                row += 1
            while ws.cell(row + 1, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row + 2)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, 0)

        all_sv_cnr = {}
        ubx_sv_cnr = {}
        time_lst = []
        diff_cnr = {}
        diff_time = {}
        per_sec_cnr_mean = []
        sec = 0
        for per_sec_info in self.all_info_list:
            valid_chl = self.all_valid_chl[sec]
            sec += 1
            if not valid_chl:
                continue
            valid_sv_id_lst = []
            cnr_8088 = {}
            cnr_ubx = {}
            try:
                per_sec_cnr = per_sec_info['cnr']
                per_sec_sv = per_sec_info['prnNOW']
            except:
                continue
            now_time = per_sec_info['chl_time']
            ubx_info = self.ubx_info_dict[now_time]
            time_lst.append(now_time)
            for chl in valid_chl:
                sv_id = per_sec_sv[chl] + 1
                valid_sv_id_lst.append(sv_id)
                cnr_8088[sv_id] = per_sec_cnr[chl]
                if sv_id in all_sv_cnr.keys():
                    all_sv_cnr[sv_id] += [per_sec_cnr[chl]]
                else:
                    all_sv_cnr[sv_id] = [per_sec_cnr[chl]]

                if sv_id > 32:
                    try:
                        cnr_ubx[sv_id] = ubx_info['5'][str(sv_id - 32)]['cnr']
                    except:
                        valid_sv_id_lst.pop()
                        continue

                    if sv_id in ubx_sv_cnr.keys():
                        ubx_sv_cnr[sv_id] += [ubx_info['5'][str(sv_id - 32)]['cnr']]
                    else:
                        ubx_sv_cnr[sv_id] = [ubx_info['5'][str(sv_id - 32)]['cnr']]
                else:
                    try:
                        cnr_ubx[sv_id] = ubx_info['0'][str(sv_id)]['cnr']
                    except:
                        valid_sv_id_lst.pop()
                        continue

                    if sv_id in ubx_sv_cnr.keys():
                        ubx_sv_cnr[sv_id] += [ubx_info['0'][str(sv_id)]['cnr']]
                    else:
                        ubx_sv_cnr[sv_id] = [ubx_info['0'][str(sv_id)]['cnr']]

            tmp_diff_sum = 0
            for sv_id in valid_sv_id_lst:
                tmp_diff = float(cnr_ubx[sv_id]) - float(cnr_8088[sv_id])
                tmp_diff_sum += tmp_diff
                if sv_id not in diff_cnr.keys():
                    diff_cnr[sv_id] = []
                    diff_time[sv_id] = []
                diff_cnr[sv_id].append(tmp_diff)
                diff_time[sv_id].append(now_time)
            tmp_len = len(valid_sv_id_lst)
            if tmp_len:
                per_sec_cnr_mean.append(tmp_diff_sum / tmp_len)
        print(
            f"mean([per_sec_cnr_mean]) = {np.mean(per_sec_cnr_mean):f}, "
            f"std([per_sec_cnr_mean]) = {np.std(per_sec_cnr_mean):f}")
        if fd_st:
            print("{:.3f}|{:.3f}".format(np.mean(per_sec_cnr_mean), np.std(per_sec_cnr_mean)), end='|', file=fd_st)
        # fig1 = plt.figure(1)
        # plt.title("compare with Ublox cnr")
        # for sv_id in diff_cnr.keys():
        #     plt.plot(diff_time[sv_id], diff_cnr[sv_id], marker='*', label='sv' + str(sv_id))
        # plt.legend()  # 不加该语句无法显示 label
        # plt.draw()
        # plt.pause(40)  # 间隔的秒数： 4s
        # plt.close(fig1)

        for key in diff_cnr.keys():
            diff_mean = round(np.mean(diff_cnr[key]), 3)
            diff_std = round(np.std(diff_cnr[key]), 3)
            our_mean = round(np.mean(all_sv_cnr[key]), 3)
            our_std = round(np.std(all_sv_cnr[key]), 3)
            ubx_mean = round(np.mean(ubx_sv_cnr[key]), 3)
            ubx_std = round(np.std(ubx_sv_cnr[key]), 3)

            write_value = [
                ['sv' + str(key), ubx_mean, ubx_std, our_mean, our_std, diff_mean, diff_std, len(diff_cnr[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def pr_dopp_union(self, aim, cmp_value=0):
        abnormal_cnt = 0
        fd_ab = open(self.path + 'chart/' + self.filename.split('.')[0] + "_abnormal_" + aim + ".txt", 'w')
        sec = 0
        time_lst = []
        # all_sv_aim = {}
        # ubx_sv_aim = {}
        diff_time = {}
        diff_aim = {}
        per_sec_diff_diff_aim_mean = []  # [mean[差值 - 差值的均值]]

        for per_sec_info in self.all_info_list:
            try:
                per_sec_sv = per_sec_info['prnNOW']
            except:
                continue
            valid_chl = self.all_valid_chl[sec]
            sec += 1
            if not valid_chl:
                continue
            now_time = per_sec_info['chl_time']
            ubx_info = self.ubx_info_dict[now_time]

            valid_sv_id_lst = []
            aim_8088 = {}
            aim_ubx = {}
            try:
                per_sec_aim = per_sec_info[aim]
            except:
                continue
            # per_sec_sv = per_sec_info['prnNOW']
            for chl in valid_chl:
                sv_id = per_sec_sv[chl] + 1
                valid_sv_id_lst.append(sv_id)
                aim_8088[sv_id] = per_sec_aim[chl]
                # if sv_id in all_sv_aim.keys():
                #     all_sv_aim[sv_id] += [per_sec_aim[chl]]
                # else:
                #     all_sv_aim[sv_id] = [per_sec_aim[chl]]

                if sv_id > 32:
                    try:
                        aim_ubx[sv_id] = ubx_info['5'][str(sv_id - 32)][aim]
                    except:
                        valid_sv_id_lst.pop()
                        continue
                    #
                    # if sv_id in ubx_sv_aim.keys():
                    #     ubx_sv_aim[sv_id] += [ubx_info['5'][str(sv_id - 32)][aim]]
                    # else:
                    #     ubx_sv_aim[sv_id] = [ubx_info['5'][str(sv_id - 32)][aim]]
                else:
                    try:
                        aim_ubx[sv_id] = ubx_info['0'][str(sv_id)][aim]
                    except:
                        valid_sv_id_lst.pop()
                        continue
                    #
                    # if sv_id in ubx_sv_aim.keys():
                    #     ubx_sv_aim[sv_id] += [ubx_info['0'][str(sv_id)][aim]]
                    # else:
                    #     ubx_sv_aim[sv_id] = [ubx_info['0'][str(sv_id)][aim]]
            tmp_len = len(valid_sv_id_lst)
            if tmp_len:
                time_lst.append(now_time)
                diff_dict = {}
                diff_list = []
                for sv_id in valid_sv_id_lst:
                    tmp_diff = round(np.fabs(float(aim_ubx[sv_id]) - float(aim_8088[sv_id])), 2)  # 差值
                    diff_dict[sv_id] = tmp_diff
                    diff_list.append(tmp_diff)

                tmp_mean = np.mean(diff_list)  # 与ubx的差 的均值
                diff_diff_mean_lst = []
                for sv_id in valid_sv_id_lst:
                    diff_diff_mean = round(np.fabs(diff_dict[sv_id] - tmp_mean), 2)
                    diff_diff_mean_lst.append(diff_diff_mean)
                    if sv_id not in diff_aim.keys():
                        diff_aim[sv_id] = []
                        diff_time[sv_id] = []
                    diff_aim[sv_id].append(diff_diff_mean)
                    diff_time[sv_id].append(now_time)
                draw_value = round(np.mean(diff_diff_mean_lst), 2)  # mean[差值 - 差值的均值]
                per_sec_diff_diff_aim_mean.append(draw_value)
                if tmp_len > 2 and cmp_value:
                    abnormal_idx = find_abnormal_data(diff_list)
                    ab_sv = valid_sv_id_lst[abnormal_idx[0]]
                    ab_diff = diff_list.pop(abnormal_idx[0])
                    tmp_mean = np.mean(diff_list)  # 与ubx的差 去除最异常的值后 的均值
                    ab_diff_diff = round(np.fabs(ab_diff - tmp_mean), 2)
                    if ab_diff_diff > cmp_value:
                        abnormal_cnt += 1
                        if draw_value > cmp_value * 10:
                            print("attention", file=fd_ab)
                        print("time=", now_time, "ab_diff_diff=", ab_diff_diff, "ab_draw=", draw_value, "ab_sv=", ab_sv,
                              "\nsv =", per_sec_sv, "\npli =", per_sec_info['pli'], "\ncnr =", per_sec_info['cnr'],
                              "\n" + aim, aim_8088, "\ndiff =", diff_dict,  # "\n", file=fd_ab)
                              "\ndiff_diff_mean =", diff_diff_mean_lst, "\n", file=fd_ab)

        return time_lst, per_sec_diff_diff_aim_mean, diff_aim, abnormal_cnt

    def pr_cmp(self, fd_st):
        head_xlsx = [['[每一秒各个卫星的PR与ublox之差 - 该秒所有卫星的PR与ublox的差值的均值] 该列表 的均值和标准差'],
                     ['sv_id', 'ave', 'std', 'cnt'], [self.filename]]
        book_name_xlsx = self.path + 'chart/' + '_compare_PR.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 3
            while ws.cell(row, 1).value:
                row += 1
            while ws.cell(row + 1, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row + 2)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, 0)

        time_lst, per_sec_diff_diff_PR_mean, diff_PR, abnormal_cnt = self.pr_dopp_union('PR', 100)
        print(
            f"mean([per_sec_diff_diff_PR_mean]) = {np.mean(per_sec_diff_diff_PR_mean):f}, "
            f"std([per_sec_diff_diff_PR_mean]) = {np.std(per_sec_diff_diff_PR_mean):f}"
            f", abnormal rate (100) = {abnormal_cnt * 100.0 / len(time_lst):f}%")
        if fd_st:
            print("{:.3f}|{:.3f}|{:.3%}".format(np.mean(per_sec_diff_diff_PR_mean),
                                                np.std(per_sec_diff_diff_PR_mean),
                                                abnormal_cnt / len(time_lst)), end='|', file=fd_st)
        fig1 = plt.figure(1)
        plt.title("compare with Ublox PR")
        plt.plot(time_lst, per_sec_diff_diff_PR_mean, marker='*', label='diff_diff_PR_mean')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_PR.png')
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

        for key in diff_PR.keys():
            diff_mean = round(np.mean(diff_PR[key]), 3)
            diff_std = round(np.std(diff_PR[key]), 3)
            write_value = [['sv' + str(key), diff_mean, diff_std, len(diff_PR[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def dopp_cmp(self, fd_st):
        head_xlsx = [['[每一秒各个卫星的DOPP与ublox之差 - 该秒所有卫星的DOPP与ublox的差值的均值] 该列表 的均值和标准差'],
                     ['sv_id', 'ave', 'std', 'cnt'], [self.filename]]
        book_name_xlsx = self.path + 'chart/' + '_compare_dopp.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 1
            while ws.cell(row, 1).value:
                row += 1
            while ws.cell(row + 1, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row + 2)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, 0)

        time_lst, per_sec_diff_diff_PR_mean, diff_PR, abnormal_cnt = self.pr_dopp_union('dopp', 5)
        print(
            f"mean([per_sec_diff_diff_dopp_mean]) = {np.mean(per_sec_diff_diff_PR_mean):f}, "
            f"std([per_sec_diff_diff_dopp_mean]) = {np.std(per_sec_diff_diff_PR_mean):f}"
            f", abnormal rate (5) = {abnormal_cnt * 100.0 / len(time_lst):f}%")
        if fd_st:
            print("{:.3f}|{:.3f}|{:.3%}".format(np.mean(per_sec_diff_diff_PR_mean),
                                                np.std(per_sec_diff_diff_PR_mean),
                                                abnormal_cnt / len(time_lst)), file=fd_st)
        fig1 = plt.figure(1)
        plt.title("compare with Ublox dopp")
        plt.plot(time_lst, per_sec_diff_diff_PR_mean, marker='*', label='diff_diff_dopp_mean')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_dopp.png')
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

        for key in diff_PR.keys():
            diff_mean = round(np.mean(diff_PR[key]), 3)
            diff_std = round(np.std(diff_PR[key]), 3)
            write_value = [['sv' + str(key), diff_mean, diff_std, len(diff_PR[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def deal_with(self):
        purpose_dict = {}
        for aim in self.purpose.keys():
            if "cnr" == aim:
                cnr_op = self.valid_chl_obj_mean_std_list(aim)
                purpose_dict[aim] = cnr_op[0]
            elif "pli" == aim:
                pli_op = self.valid_chl_obj_mean_std_list(aim)
                purpose_dict[aim] = pli_op[0]
            # elif "pos" == aim:

    def valid_chl_obj_mean_std_list(self, key):
        chl_obj_mean = []
        chl_obj_std = []
        for sec, obj_lst in enumerate(self.all_info_dict[key]):
            valid_obj = []
            for i in self.all_valid_chl[sec]:
                valid_obj.append(obj_lst[i])

            if valid_obj:
                chl_obj_mean.append(np.mean(valid_obj))
                chl_obj_std.append(np.std(valid_obj))
            else:
                chl_obj_mean.append(0)
                chl_obj_std.append(0)
        return chl_obj_mean, chl_obj_std

    def pli_PR(self):
        PR_diff_diff_pli_mean = {}
        PR_diff_diff_dli_mean = {}
        PR_diff_diff_dict_pli = {}
        PR_diff_diff_dict_dli = {}

        for per_sec_info in self.all_info_list:
            now_time = per_sec_info['chl_time']
            ubx_info = self.ubx_info_dict[now_time]
            try:
                per_sec_sv = per_sec_info['prnNOW']
            except:
                continue
            dli_lst = per_sec_info["dli"]
            pli_lst = per_sec_info["pli"]
            PR_lst = per_sec_info["PR"]
            chl_valid = []
            for idx, pli in enumerate(pli_lst):
                if pli != 100:
                    chl_valid.append(idx)
            valid_sv = []
            dli_8088 = {}
            pli_8088 = {}
            PR_8088 = {}
            PR_ubx = {}

            for idx in chl_valid:
                sv_id = per_sec_sv[idx] + 1
                valid_sv.append(sv_id)

                PR_8088[sv_id] = PR_lst[idx]
                pli_8088[sv_id] = pli_lst[idx]
                dli_8088[sv_id] = dli_lst[idx]
                if sv_id > 32:
                    try:
                        PR_ubx[sv_id] = ubx_info['5'][str(sv_id - 32)]["PR"]
                    except:
                        valid_sv.pop()
                else:
                    try:
                        PR_ubx[sv_id] = ubx_info['0'][str(sv_id)]["PR"]
                    except:
                        valid_sv.pop()

            tmp_len = len(valid_sv)
            if tmp_len:
                diff_dict = {}
                diff_list = []
                for sv_id in valid_sv:
                    tmp_diff = round(np.fabs(float(PR_ubx[sv_id]) - float(PR_8088[sv_id])), 2)  # 差值
                    diff_dict[sv_id] = tmp_diff
                    diff_list.append(tmp_diff)

                mean_val = self.del_ab_val_calc_mean(diff_list)

                for key in diff_dict.keys():
                    diff_diff_mean = np.abs(diff_dict[key] - mean_val)
                    if diff_diff_mean > 500 or diff_diff_mean < -500:
                        continue
                    if pli_8088[key] in PR_diff_diff_dict_pli.keys():
                        PR_diff_diff_dict_pli[pli_8088[key]] += [diff_diff_mean, ]
                    else:
                        PR_diff_diff_dict_pli[pli_8088[key]] = [diff_diff_mean, ]

                    if dli_8088[key] in PR_diff_diff_dict_dli.keys():
                        PR_diff_diff_dict_dli[dli_8088[key]] += [diff_diff_mean, ]
                    else:
                        PR_diff_diff_dict_dli[dli_8088[key]] = [diff_diff_mean, ]

        for key in PR_diff_diff_dict_pli.keys():
            PR_diff_diff_pli_mean[key] = np.mean(PR_diff_diff_dict_pli[key])
        for key in PR_diff_diff_dict_dli.keys():
            PR_diff_diff_dli_mean[key] = np.mean(PR_diff_diff_dict_dli[key])

        return PR_diff_diff_pli_mean, PR_diff_diff_dli_mean

    def del_ab_val_calc_mean(self, arr: list):
        """
        ram arr:  an array
        :return:  the mean of arr without the abnormal items
        """
        if len(arr) < 3:
            return np.mean(arr)
        max_val = max(arr)
        min_val = min(arr)

        max_idx = arr.index(max_val)
        min_idx = arr.index(min_val)

        del_max = arr.copy()
        del_max.pop(max_idx)
        del_min = arr.copy()
        del_min.pop(min_idx)

        del_max_std = np.std(del_max)
        del_min_std = np.std(del_min)

        if del_max_std < del_min_std:
            diff_ = np.fabs(max_val - np.mean(del_max))
            if diff_ > 100:
                return self.del_ab_val_calc_mean(del_max)
            else:
                return np.mean(arr)
        else:
            diff_ = np.fabs(min_val - np.mean(del_min))
            if diff_ > 100:
                return self.del_ab_val_calc_mean(del_min)
            else:
                return np.mean(arr)

    @staticmethod
    def num_of_valid_pli(arr):
        n = 0
        for i in arr:
            if i != 100:
                n += 1
        return n

    @staticmethod
    def valid_pli_and_index(arr):
        valid_pli, idx = [], []
        for i, item in enumerate(arr):
            if item != 100:
                valid_pli.append(item)
                idx.append(i)
        return valid_pli, idx

    def not_fix_analysis(self):
        reason = {"sv_less_4": [], "pli_too_bad": [], "unknown": []}
        for info in self.all_info_list:
            GGA = info["GGA"]
            time = info["chl_time"]
            if GGA['valid'] == 0:
                pli = info["pli"]
                try:
                    if info["val_num"] < 4:
                        reason["sv_less_4"].append(time)
                        continue
                except:
                    reason["sv_less_4"].append(time)
                    continue

                valid_pli, _o_ = self.valid_pli_and_index(pli)
                if np.mean(valid_pli) > 20:
                    # print("pli too bad")
                    reason["pli_too_bad"].append(time)
                    continue
                reason["unknown"].append(time)
        print("sv_less_4: ", len(reason["sv_less_4"]))
        print("pli_too_bad: ", len(reason["pli_too_bad"]))
        print("unknown: ", len(reason["unknown"]))
