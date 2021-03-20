import time
import os
import sys
import re
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from base_function import ecef_to_enu, lla_to_xyz, calc_True_Txyz, write_excel_xlsx,\
    find_abnormal_data, xyz_to_lla, degree_to_dms

# from typing import List, Iterator, Union,

__all__ = ['DrawPicture', 'LogAnalysis', 'LogParser']


class DrawPicture:
    def __init__(self, _path_, _file_):
        self.path = _path_
        self.file = _file_

    def draw_one_picture(self, x_axis, y_axis_dict, title=''):
        fig1 = plt.figure(1)
        plt.title(self.file + title)
        plt.xlabel("sec of week")
        for key in y_axis_dict:
            plt.plot(x_axis, y_axis_dict[key], marker='o', linestyle=':', label=key)
        plt.legend()    # 不加该语句无法显示 label
        plt.draw()
        plt.pause(5)
        plt.close(fig1)


class LogAnalysis:
    def __init__(self, target_file, __purpose__, ubx_file, fd_st):
        self.valid_chl_flag = [0, ]
        self.cmp_enable = 0
        self.cmp_support = 0

        self.fd_st = fd_st

        if len(ubx_file):   # 有U-blox的文件
            self.cmp_support = 1
            self.f_ubx = open(ubx_file, 'r', errors="ignore")

        if isinstance(__purpose__, dict) and len(target_file):
            self.filename = target_file.split('/')[-1]
            self.path = target_file.split(self.filename)[0]
            self.f_our = open(target_file, 'r', errors="ignore")
            self.purpose = __purpose__.copy()
            _target_row = ["chl_time", ]
            for key in __purpose__.keys():
                if key in LogParser.purpose_need_row.keys():
                    _target_row += LogParser.purpose_need_row[key]
                else:
                    del self.purpose[key]
                    print("can't support the <%s> purpose" % key)

                if key in ["cnr", "pos", "pr", "dopp", "sv_keep"]:
                    self.cmp_enable = 1
            self.target_row = list(set(_target_row))
        else:
            sys.exit("you must input our path file, "
                     "1 dict, the key is target_row and value are you want operate\n"
                     "the ubx path file is not essential")
        self.all_info_list: list = []   # [{"cnr": [], "pli": [], "pr": [], ..}, {2sec info}, ... ]
        self.all_info_dict: dict = {}   # {"cnr": [[1sec], [2sec], ... ], "pli":[[1sec], [2sec], ... ],  ..}
        self.all_valid_chl: list = []
        self.ubx_info_dict: dict = {}   # {time: [each sv_info], time: [each sv_info], }
        self.ubx_bak_time = 0
        self.final_pos_str = ''

    def final_pos_analysis(self, true_xyz):
        original_str = self.final_pos_str
        print(original_str)
        str_xyz = original_str[original_str.index('xyz,') + 5: original_str.index(', idx')]
        str_xyz = str_xyz.split(',')
        final_xyz = list([float(x) for x in str_xyz])
        final_lla = xyz_to_lla(final_xyz[0], final_xyz[1], final_xyz[2])
        final_enu = ecef_to_enu(true_xyz[0], true_xyz[1], true_xyz[2], final_lla[0], final_lla[1], final_lla[2])
        print("\n8088 最终定位位置:", final_lla)
        final_lat = degree_to_dms(str(final_lla[0]))[0]
        final_lon = degree_to_dms(str(final_lla[1]))[0]
        print(final_lat, final_lon, final_lla[2])
        print("东北天 = ", final_enu)

        dis_en = np.linalg.norm(np.array([final_enu[0], final_enu[1]]))
        dis_T_xyz = np.linalg.norm(np.array(final_xyz) - np.array(true_xyz))
        print("平面误差 dis_en = ", round(dis_en, 3))
        print("与真值的距离 dis_T_xyz = ", round(dis_T_xyz, 3))
        if self.fd_st:
            print("{:.3f}|{:.3f}".format(dis_T_xyz, dis_en), end='|', file=self.fd_st)

    def pli_abnormal_pli_mean_cnr_mean(self):
        _fd_ = open(self.path + 'chart/' + self.filename.split('.')[0] + '_ab_pli.txt', 'w')
        pli_mean = []
        cnr_mean = []
        time_lst = []
        pli_row_bak = []
        sum_bak = 100
        abnormal_pli_cnt = 0
        for per_sec_info in self.all_info_list:
            pli = []
            idx = []
            try:
                _tmp_pli_lst = per_sec_info['pli']
            except:
                continue
            for i, item in enumerate(_tmp_pli_lst):
                if item != 100:
                    pli.append(item)
                    idx.append(i)
            if len(pli):
                pli_mean.append(np.mean(np.array(pli)))
                pli_sum = np.sum(pli)
                if pli_sum / sum_bak > 2 and pli_mean[-1] > 15 and len(pli) > 2:
                    print(pli_row_bak, per_sec_info['chl_time'], per_sec_info['pli'], file=_fd_)
                    abnormal_pli_cnt += 1
                sum_bak = pli_sum
            else:
                sum_bak = 100
                pli_mean.append(0)
            pli_row_bak = per_sec_info['pli']

            if idx:
                cnr = []
                for i in idx:
                    cnr.append(per_sec_info['cnr'][i])
                cnr_mean.append(np.mean(np.array(cnr)))
            else:
                cnr_mean.append(60)

            time_lst.append(per_sec_info['chl_time'])

        _fd_.close()
        print("abnormal pli rate = {:.2%}".format(abnormal_pli_cnt / len(pli_mean)))
        print("mean(pli_mean) = {:.2f}, std(pli_mean) = {:.2f}".format(np.mean(pli_mean), np.std(pli_mean)))
        print("mean(cnr_mean) = {:.2f}, std(cnr_mean) = {:.2f}".format(np.mean(cnr_mean), np.std(cnr_mean)))
        if self.fd_st:
            print("{:.3f}|{:.3f}|{:.3%}|{:.3f}|{:.3f}".format(
                np.mean(pli_mean), np.std(pli_mean), abnormal_pli_cnt / len(pli_mean),
                np.mean(cnr_mean), np.std(cnr_mean)), end='|', file=self.fd_st)

        fig1 = plt.figure(1)
        plt.title("per sec pli list's mean and cnr list's mean")
        plt.plot(time_lst, pli_mean, marker='*', label='pli')
        plt.plot(time_lst, cnr_mean, marker='o', label='cnr')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + "chart/" + self.filename[:-4] + "_per_sec_pli_cnr.png")
        # plt.show()
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

    def static_pos_cmp(self, true_xyz):
        RMC_list = self.all_info_dict['RMC']
        time_list = self.all_info_dict['chl_time']
        all_diff_xyz = []
        all_EN = []
        time_pos = []
        diff_vel = []
        time_vel = []
        gga_1_rmc_N = 0
        sec = 0
        for info in self.all_info_dict['GGA']:
            rmc_dict = RMC_list[sec]
            SoW = time_list[sec]        # second of week

            if info['valid']:
                diff_xyz = (np.array(true_xyz) - np.array(info['xyz']))
                all_diff_xyz.append(diff_xyz)  # 收集点(x,y,z)3个轴的误差
                ENU = ecef_to_enu(true_xyz[0], true_xyz[1], true_xyz[2], info['lat'], info['lon'], info['high'])
                all_EN.append(ENU[:2])  # 收集东北
                time_pos.append(SoW)

            if rmc_dict['valid']:
                diff_vel.append(rmc_dict['speed'])
                time_vel.append(SoW)
            else:
                if info['valid']:
                    gga_1_rmc_N += 1
            sec += 1

        all_dis_xyz = [np.linalg.norm(diff_xyz) for diff_xyz in all_diff_xyz]
        all_dis_EN = [np.linalg.norm(en) for en in all_EN]
        len_tmp = len(time_list)
        print("fix rate = {:.3%}, warning rate = {:.3%}".format(len(time_pos) / len_tmp, gga_1_rmc_N / len_tmp))
        if self.fd_st:
            self.sort_and_print_50_95_99(all_dis_xyz, "all_dis_xyz")
            self.sort_and_print_50_95_99(all_dis_EN, "all_dis_EN")
            self.sort_and_print_50_95_99(diff_vel, "diff_vel")
            print("{:.3%}|{:.3%}".format(len(time_pos) / len_tmp, gga_1_rmc_N / len_tmp),
                  end='|', file=self.fd_st)

        fig1 = plt.figure(1)
        plt.title("distance with true position")
        plt.plot(time_pos, all_dis_xyz, marker='*', label='dis_xyz')
        plt.plot(time_pos, all_dis_EN, marker='o', label='dis_EN')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_pos.png')
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

    def sort_and_print_50_95_99(self, aim_list, keyword):
        sort_list = np.sort(aim_list)
        len_tmp = len(sort_list)
        percentage_50 = sort_list[int(len_tmp * 0.5)]
        percentage_95 = sort_list[int(len_tmp * 0.95)]
        percentage_99 = sort_list[int(len_tmp * 0.99)]
        std_list = np.std(sort_list)
        print(keyword + " 50% = {:f}, 95% = {:f}, 99% = {:f}, std = {:f}"
              .format(percentage_50, percentage_95, percentage_99, std_list))
        print("{:.3f}|{:.3f}|{:.3f}|{:.3f}".format(percentage_50, percentage_95, percentage_99, std_list),
              end='|', file=self.fd_st)

    def cnr_cmp(self):
        head_xlsx = [['', "[ublox每一秒的CNR]该列表的均值和标准差", '', "[8088每一秒的CNR]该列表的均值和标准差", '',
                      "[每一秒各个卫星的载噪比与ublox的载噪比之差] 该列表 的均值和标准差"],
                     ['', 'ubx_ave', 'ubx_std', 'Our_ave', 'Our_std', 'diff_ave', 'diff_std', 'cnt'],
                     [self.filename, ]]
        book_name_xlsx = self.path + 'chart/' + '_compare_cnr.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 3
            while ws.cell(row, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row+1)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, row_cnt=0)

        all_sv_cnr = {}
        ubx_sv_cnr = {}
        time_lst = []
        diff_cnr = {}
        diff_time = {}
        per_sec_cnr_mean = []
        sec = 0
        for per_sec_info in self.all_info_list:
            valid_chl = self.all_valid_chl[sec]
            sec += 1
            if not valid_chl:
                continue
            valid_sv_id_lst = []
            cnr_8088 = {}
            cnr_ubx = {}
            try:
                per_sec_cnr = per_sec_info['cnr']
                per_sec_sv = per_sec_info['prnNOW']
            except:
                continue
            now_time = per_sec_info['chl_time']
            ubx_info = self.ubx_info_dict[now_time]
            time_lst.append(now_time)
            for chl in valid_chl:
                sv_id = per_sec_sv[chl] + 1
                valid_sv_id_lst.append(sv_id)
                cnr_8088[sv_id] = per_sec_cnr[chl]
                if sv_id in all_sv_cnr.keys():
                    all_sv_cnr[sv_id] += [per_sec_cnr[chl]]
                else:
                    all_sv_cnr[sv_id] = [per_sec_cnr[chl]]

                if sv_id > 32:
                    try:
                        cnr_ubx[sv_id] = ubx_info['5'][str(sv_id - 32)]['cnr']
                    except:
                        valid_sv_id_lst.pop()
                        continue

                    if sv_id in ubx_sv_cnr.keys():
                        ubx_sv_cnr[sv_id] += [ubx_info['5'][str(sv_id - 32)]['cnr']]
                    else:
                        ubx_sv_cnr[sv_id] = [ubx_info['5'][str(sv_id - 32)]['cnr']]
                else:
                    try:
                        cnr_ubx[sv_id] = ubx_info['0'][str(sv_id)]['cnr']
                    except:
                        valid_sv_id_lst.pop()
                        continue

                    if sv_id in ubx_sv_cnr.keys():
                        ubx_sv_cnr[sv_id] += [ubx_info['0'][str(sv_id)]['cnr']]
                    else:
                        ubx_sv_cnr[sv_id] = [ubx_info['0'][str(sv_id)]['cnr']]

            tmp_diff_sum = 0
            for sv_id in valid_sv_id_lst:
                tmp_diff = float(cnr_ubx[sv_id]) - float(cnr_8088[sv_id])
                tmp_diff_sum += tmp_diff
                if sv_id not in diff_cnr.keys():
                    diff_cnr[sv_id] = []
                    diff_time[sv_id] = []
                diff_cnr[sv_id].append(tmp_diff)
                diff_time[sv_id].append(now_time)
            tmp_len = len(valid_sv_id_lst)
            if tmp_len:
                per_sec_cnr_mean.append(tmp_diff_sum/tmp_len)
        print(
            f"mean([per_sec_cnr_mean]) = {np.mean(per_sec_cnr_mean):f}, "
            f"std([per_sec_cnr_mean]) = {np.std(per_sec_cnr_mean):f}")
        if self.fd_st:
            print("{:.3f}|{:.3f}".format(np.mean(per_sec_cnr_mean), np.std(per_sec_cnr_mean)), end='|', file=self.fd_st)
        # fig1 = plt.figure(1)
        # plt.title("compare with Ublox cnr")
        # for sv_id in diff_cnr.keys():
        #     plt.plot(diff_time[sv_id], diff_cnr[sv_id], marker='*', label='sv' + str(sv_id))
        # plt.legend()  # 不加该语句无法显示 label
        # plt.draw()
        # plt.pause(40)  # 间隔的秒数： 4s
        # plt.close(fig1)

        for key in diff_cnr.keys():
            diff_mean = round(np.mean(diff_cnr[key]), 3)
            diff_std = round(np.std(diff_cnr[key]), 3)
            our_mean = round(np.mean(all_sv_cnr[key]), 3)
            our_std = round(np.std(all_sv_cnr[key]), 3)
            ubx_mean = round(np.mean(ubx_sv_cnr[key]), 3)
            ubx_std = round(np.std(ubx_sv_cnr[key]), 3)

            write_value = [['sv' + str(key), ubx_mean, ubx_std, our_mean, our_std, diff_mean, diff_std, len(diff_cnr[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def pr_dopp_union(self, aim, ab_value=0):
        abnormal_cnt = 0
        fd_ab = open(self.path + 'chart/' + self.filename.split('.')[0] + "_abnormal_" + aim + ".txt", 'w')
        sec = 0
        time_lst = []
        # all_sv_aim = {}
        # ubx_sv_aim = {}
        diff_time = {}
        diff_aim = {}
        per_sec_diff_diff_aim_mean = []     # [mean[差值 - 差值的均值]]

        for per_sec_info in self.all_info_list:
            valid_chl = self.all_valid_chl[sec]
            sec += 1
            if not valid_chl:
                continue
            now_time = per_sec_info['chl_time']
            ubx_info = self.ubx_info_dict[now_time]

            valid_sv_id_lst = []
            aim_8088 = {}
            aim_ubx = {}
            try:
                per_sec_aim = per_sec_info[aim]
            except:
                continue
            per_sec_sv = per_sec_info['prnNOW']
            for chl in valid_chl:
                sv_id = per_sec_sv[chl] + 1
                valid_sv_id_lst.append(sv_id)
                aim_8088[sv_id] = per_sec_aim[chl]
                # if sv_id in all_sv_aim.keys():
                #     all_sv_aim[sv_id] += [per_sec_aim[chl]]
                # else:
                #     all_sv_aim[sv_id] = [per_sec_aim[chl]]

                if sv_id > 32:
                    try:
                        aim_ubx[sv_id] = ubx_info['5'][str(sv_id - 32)][aim]
                    except:
                        valid_sv_id_lst.pop()
                        continue
                    #
                    # if sv_id in ubx_sv_aim.keys():
                    #     ubx_sv_aim[sv_id] += [ubx_info['5'][str(sv_id - 32)][aim]]
                    # else:
                    #     ubx_sv_aim[sv_id] = [ubx_info['5'][str(sv_id - 32)][aim]]
                else:
                    try:
                        aim_ubx[sv_id] = ubx_info['0'][str(sv_id)][aim]
                    except:
                        valid_sv_id_lst.pop()
                        continue
                    #
                    # if sv_id in ubx_sv_aim.keys():
                    #     ubx_sv_aim[sv_id] += [ubx_info['0'][str(sv_id)][aim]]
                    # else:
                    #     ubx_sv_aim[sv_id] = [ubx_info['0'][str(sv_id)][aim]]
            tmp_len = len(valid_sv_id_lst)
            if tmp_len:
                time_lst.append(now_time)
                diff_dict = {}
                diff_list = []
                for sv_id in valid_sv_id_lst:
                    tmp_diff = np.fabs(float(aim_ubx[sv_id]) - float(aim_8088[sv_id]))  # 差值

                    diff_dict[sv_id] = tmp_diff
                    diff_list.append(tmp_diff)
                tmp_mean = np.mean(diff_list)     # 与ubx的差 的均值
                diff_diff_mean_lst = []
                for sv_id in valid_sv_id_lst:
                    diff_diff_mean = np.fabs(diff_dict[sv_id] - tmp_mean)
                    diff_diff_mean_lst.append(diff_diff_mean)
                    if sv_id not in diff_aim.keys():
                        diff_aim[sv_id] = []
                        diff_time[sv_id] = []
                    diff_aim[sv_id].append(diff_diff_mean)
                    diff_time[sv_id].append(now_time)
                per_sec_diff_diff_aim_mean.append(np.mean(diff_diff_mean_lst))   # [mean[差值 - 差值的均值]]
                if tmp_len > 2 and ab_value:
                    abnormal_idx = find_abnormal_data(diff_diff_mean_lst)
                    abnormal_diff = diff_diff_mean_lst[abnormal_idx[0]]
                    if abnormal_diff > ab_value:
                        abnormal_cnt += 1
                        print("abnormal_sv =", valid_sv_id_lst[abnormal_idx[0]], "time =", now_time,
                              "\nsv =", per_sec_sv, "\npli =", per_sec_info['pli'], "\ncnr =", per_sec_info['cnr'],
                              "\n" + aim, aim_8088, "\ndiff =", diff_dict,
                              "\ndiff_diff_mean =", diff_diff_mean_lst, file=fd_ab)

        return time_lst, per_sec_diff_diff_aim_mean, diff_aim, abnormal_cnt

    def pr_cmp(self):
        head_xlsx = [['[每一秒各个卫星的PR与ublox之差 - 该秒所有卫星的PR与ublox的差值的均值] 该列表 的均值和标准差'],
                     ['sv_id', 'ave', 'std', 'cnt'], [self.filename]]
        book_name_xlsx = self.path + 'chart/' + '_compare_PR.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 3
            while ws.cell(row, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row+1)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, row_cnt=0)

        time_lst, per_sec_diff_diff_PR_mean, diff_PR, abnormal_cnt = self.pr_dopp_union('PR', 100)
        print(
            f"mean([per_sec_diff_diff_PR_mean]) = {np.mean(per_sec_diff_diff_PR_mean):f}, "
            f"std([per_sec_diff_diff_PR_mean]) = {np.std(per_sec_diff_diff_PR_mean):f}"
            f", abnormal rate (100) = {abnormal_cnt*100.0/len(time_lst):f}%")
        if self.fd_st:
            print("{:.3f}|{:.3f}|{:.3%}".format(np.mean(per_sec_diff_diff_PR_mean),
                                         np.std(per_sec_diff_diff_PR_mean),
                                         abnormal_cnt/len(time_lst)), end='|', file=self.fd_st)
        fig1 = plt.figure(1)
        plt.title("compare with Ublox PR")
        plt.plot(time_lst, per_sec_diff_diff_PR_mean, marker='*', label='diff_diff_PR_mean')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_PR.png')
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

        for key in diff_PR.keys():
            diff_mean = round(np.mean(diff_PR[key]), 3)
            diff_std = round(np.std(diff_PR[key]), 3)
            write_value = [['sv' + str(key), diff_mean, diff_std, len(diff_PR[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def dopp_cmp(self):
        head_xlsx = [['[每一秒各个卫星的DOPP与ublox之差 - 该秒所有卫星的DOPP与ublox的差值的均值] 该列表 的均值和标准差'],
                     ['sv_id', 'ave', 'std', 'cnt'], [self.filename]]
        book_name_xlsx = self.path + 'chart/' + '_compare_dopp.xlsx'

        try:
            wb = openpyxl.load_workbook(book_name_xlsx)
            ws = wb.active
            row = 1
            while ws.cell(row, 1).value:
                row += 1
            row_xlsx = write_excel_xlsx(ws, [[self.filename]], row + 1)
        except:
            wb = openpyxl.Workbook()  # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
            ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
            row_xlsx = write_excel_xlsx(ws, head_xlsx, row_cnt=0)

        time_lst, per_sec_diff_diff_PR_mean, diff_PR, abnormal_cnt = self.pr_dopp_union('dopp', 5)
        print(
            f"mean([per_sec_diff_diff_dopp_mean]) = {np.mean(per_sec_diff_diff_PR_mean):f}, "
            f"std([per_sec_diff_diff_dopp_mean]) = {np.std(per_sec_diff_diff_PR_mean):f}"
            f", abnormal rate (5) = {abnormal_cnt*100.0/len(time_lst):f}%")
        if self.fd_st:
            print("{:.3f}|{:.3f}|{:.3%}".format(np.mean(per_sec_diff_diff_PR_mean),
                                                np.std(per_sec_diff_diff_PR_mean),
                                                abnormal_cnt/len(time_lst)), file=self.fd_st)
        fig1 = plt.figure(1)
        plt.title("compare with Ublox dopp")
        plt.plot(time_lst, per_sec_diff_diff_PR_mean, marker='*', label='diff_diff_dopp_mean')
        plt.legend()  # 不加该语句无法显示 label
        plt.draw()
        plt.savefig(self.path + 'chart/' + self.filename[:-4] + '_cmp_dopp.png')
        plt.pause(4)  # 间隔的秒数： 4s
        plt.close(fig1)

        for key in diff_PR.keys():
            diff_mean = round(np.mean(diff_PR[key]), 3)
            diff_std = round(np.std(diff_PR[key]), 3)
            write_value = [['sv' + str(key), diff_mean, diff_std, len(diff_PR[key])]]
            row_xlsx = write_excel_xlsx(ws, write_value, row_xlsx)
        wb.save(book_name_xlsx)
        wb.close()

    def deal_with(self):
        purpose_dict = {}
        for aim in self.purpose.keys():
            if "cnr" == aim:
                cnr_op = self.valid_chl_obj_mean_std_list(aim)
                purpose_dict[aim] = cnr_op[0]
            elif "pli" == aim:
                pli_op = self.valid_chl_obj_mean_std_list(aim)
                purpose_dict[aim] = pli_op[0]
            # elif "pos" == aim:

    def valid_chl_obj_mean_std_list(self, key):
        chl_obj_mean = []
        chl_obj_std = []
        for sec, obj_lst in enumerate(self.all_info_dict[key]):
            valid_obj = []
            for i in self.all_valid_chl[sec]:
                valid_obj.append(obj_lst[i])

            if valid_obj:
                chl_obj_mean.append(np.mean(valid_obj))
                chl_obj_std.append(np.std(valid_obj))
            else:
                chl_obj_mean.append(0)
                chl_obj_std.append(0)
        return chl_obj_mean, chl_obj_std

    def valid_chl_list(self):
        valid_chl_lst = []
        for sv_info in self.all_info_dict["svINFO"]:
            valid_chl_lst.append(self.valid_chl_per_sec(sv_info))
        return valid_chl_lst

    def valid_chl_per_sec(self, sv_info):
        valid_chl = []
        for idx, item in enumerate(sv_info):
            if item in self.valid_chl_flag:
                valid_chl.append(idx)
        return valid_chl


class LogParser(LogAnalysis):
    purpose_need_row = {"cnr": ["cnr", "svINFO", "prnNOW"], "pli": ["pli", "svINFO", "prnNOW"],
                        "pos": ["GGA", "RMC"], "PR": ["PR", "svINFO", "prnNOW"], "dopp": ["dopp", "svINFO", "prnNOW"]}
    row_flag_dict = {"cnr": ['cnr:', ''], "pli": ['pli ', ''], "svINFO": ['SV INFO', ''], "prnNOW": ['PRN NOW', ''],
                     "RMC": ['$GPRMC,', ''], "GGA": ['$GPGGA,', ',*'], "GFM": ['$GPGFM', ''],
                     "chl_time": ['CHL TIME,', ''],
                     "dopp": ['CHL DOPP,', 'ir'], 'PR': ['CHL PR,', 'ir'], "fix_sv": ['PV, val num', '']}
    NUM_COMMA = {"GGA": 14, "RMC": 12, "GFM": 14,
                 "chl_time": 3, "dopp": 11, "PR": 10,
                 "svINFO": 9, "prnNOW": 9, "fix_sv": 2,
                 "cnr": 9, "pli": 9}

    def __init__(self, target_file: str, __purpose__: dict, ubx_file='', fd_st=''):
        super().__init__(target_file, __purpose__, ubx_file, fd_st)
        self.row_cnt = 0
        self.restart = 0
        self.time_bak = -1
        self.pos = 0
        self.FILE_LEN = self.f_our.seek(0, 2)  # 把文件指针移到文件末尾并移动0
        self.f_our.seek(0, 0)  # 把文件指针移到文件起始并移动0
        print(self.path, self.filename)
        print("目的:", self.purpose, "\n需要解析的行:", self.target_row)

    def is_intact(self, string, key_word):
        """
        :param key_word: 语句关键字
        :param string: 要检测的字符串
        """
        if string.count(',') == self.NUM_COMMA[key_word]:
            return True
        else:
            return False

    def one_second_field(self, flag_one_sec="bit lck"):
        one_sec_row = {}
        line = self.f_our.readline()
        self.row_cnt += 1
        while line:
            line = self.f_our.readline()
            self.row_cnt += 1

            if "tot cnt:" in line:
                self.final_pos_str = line
            elif "set time:" in line:
                if self.restart == 2:
                    self.restart = 1

            for target in self.target_row:
                if line.startswith(self.row_flag_dict[target][0]) and self.row_flag_dict[target][1] in line \
                        and target not in one_sec_row.keys():
                    if self.is_intact(line, target):
                        one_sec_row[target] = line
                    else:
                        print("The  %d row <<%s>> is incomplete" % (self.row_cnt, line))
            ''' 重启标志 '''
            if "cce load over" in line:
                self.restart = 2
                self.pos = self.f_our.tell()
                return {}
            ''' 1 秒标志 '''
            if flag_one_sec in line:
                self.pos = self.f_our.tell()
                return one_sec_row

        self.pos = self.f_our.tell()
        return one_sec_row

    def get_ubx_time(self):
        line = self.f_ubx.readline()
        while line:
            if "gpsTOW" in line:
                ret = re.findall(r"\d+", line)
                ubx_time = round(int(ret[1]) / 1000)
                if self.ubx_bak_time - ubx_time > 30240:
                    ubx_time += 604800  # 到下一周了
                self.ubx_bak_time = ubx_time
                return ubx_time
            line = self.f_ubx.readline()
        return -1

    def parser_ubx_txt(self):
        next_field = 0
        all_sv_info_dict = {}   # {gps: {'svId': {'cnr': 38, 'pli': 5, }, }, BD: {'svId': {'cnr': 45, 'pli': 3, }, } }
        line = self.f_ubx.readline()
        while line:
            line = self.f_ubx.readline()
            if not line:
                print("ubx file end")
                return {}
            if "--------" in line:
                break
            elif "rcvTow" in line:
                next_field = 1
                continue

            ret = line.split()
            if not bool(ret) or ret[0].isalpha():   # ret = [] or ret[0] is not num
                continue

            if next_field:
                if float(ret[0]) < 1e6:
                    continue
                if ret[3] not in ['0', '5']:
                    continue
                if ret[4] not in all_sv_info_dict[ret[3]]:
                    continue

                all_sv_info_dict[ret[3]][ret[4]]["PR"] = float(ret[0])

                continue

            if ret[0].isdigit():    # 字符都是数字，为真返回 Ture，否则返回 False
                if ret[0] not in ['0', '5']:
                    continue
                if ret[0] in all_sv_info_dict.keys():
                    all_sv_info_dict[ret[0]][ret[1]] = {"cnr": int(ret[2]), "dopp": float(ret[5])}
                else:
                    all_sv_info_dict[ret[0]] = {ret[1]: {}}
                    all_sv_info_dict[ret[0]][ret[1]] = {"cnr": int(ret[2]), "dopp": float(ret[5])}
        return all_sv_info_dict

    def parser_file(self):
        _all_info_list = []
        t_our = -1
        t_ubx = -1
        val_our = 0
        val_ubx = 0
        max_td = 0.2            # unit s
        while self.pos < self.FILE_LEN:
            if self.cmp_enable and self.cmp_support:
                if not val_our:
                    field_dict = self.one_second_field()
                    if not bool(field_dict):    # this second field is empty
                        continue
                    parser_field_dict = self.parser_field(field_dict)
                    if self.restart:
                        continue
                    try:
                        t_our = parser_field_dict["chl_time"]
                    except:
                        continue
                    if t_our == -1:
                        continue
                #    # get_time_anchor(file, f_ptr)
                # else:
                #     t_our = get_chl_time(file, f_ptr)
                #     val_our = is_val_time_our(t_our)

                if not val_ubx:
                    t_ubx = self.get_ubx_time()
                    if t_ubx == -1:
                        print("ubx file end, can't match the time ", t_our)
                        break

                td = t_our - t_ubx
                if td > max_td:
                    val_our = 1

                elif td < -max_td:
                    val_ubx = 1

                else:
                    val_our = 1
                    val_ubx = 1
                if val_our and val_ubx:     # time match!
                    val_our = 0
                    val_ubx = 0
                    tmp_dict = self.parser_ubx_txt()
                    if not bool(tmp_dict):
                        continue
                    self.ubx_info_dict[t_our] = tmp_dict
                    _all_info_list.append(parser_field_dict)
            else:
                field_dict = self.one_second_field()
                if bool(field_dict):
                    file_dict = self.parser_field(field_dict)
                    try:
                        if file_dict["chl_time"] == -1:
                            continue
                    except:
                        continue
                    _all_info_list.append(file_dict)

            #     append_our_info()
            #     line = self.f_ubx.readline()
            #     while line:
            #         # ubx version
            #         field_dict = self.one_second_field()
            #         if not val_ubx:
            #             get_time_anchor(file, f_ptr)
            #         elif val_ubx:
            #             t_our = get_chl_time(file, f_ptr)
            #             val_ubx = is_val_time_ubx(t_our)
            #         if val_our and val_ubx:
            #             td = t_our - t_ubx
            #         # whether to continue our file
            #         if not val_ubx or td > max_td:
            #             continue
            #         # time match!
            #         elif (td < max_td and td > -max_td)
            #             append_val_info()
            #             break
            #             append_ubx_info()
            #
            #
            #
            # if bool(field_dict):
            #     file_dict = self.parser_field(field_dict)
            #     if file_dict["chl_time"] == -1:
            #         continue
            #     _all_info_list.append(file_dict)
            #
            #
            #
            #     if self.cmp_enable and self.cmp_support:
            #         if (file_dict["chl_time"] - self.ubx_bak_time) > 0.2:
            #             tmp_field_flag = 0
            #             line = self.f_ubx.readline()
            #             while line:
            #                 if "-----------" in line:
            #                     break
            #                 elif "rcvTow" in line:
            #                     tmp_field_flag = 1
            #                 else:
        # while self.pos < self.FILE_LEN:
        #     field_dict = self.one_second_field()
        #     if bool(field_dict):
        #         file_dict = self.parser_field(field_dict)
        #         if file_dict["chl_time"] == -1:
        #             continue
        #         _all_info_list.append(file_dict)
        #
        #
        #
        #         if self.cmp_enable and self.cmp_support:
        #             if (file_dict["chl_time"] - self.ubx_bak_time) > 0.2:
        #                 tmp_field_flag = 0
        #                 line = self.f_ubx.readline()
        #                 while line:
        #                     if "-----------" in line:
        #                         break
        #                     elif "rcvTow" in line:
        #                         tmp_field_flag = 1
        #                     else:

        if len(_all_info_list) == 0:
            sys.exit("find nothing valid information. Please make sure the input file is right")
        self.all_info_list = _all_info_list  # [{"cnr": [], "pli": [], "pr": [], }, {2sec info}, ... ]
        self._transpose_()
        self.all_valid_chl = self.valid_chl_list()

    def parser_field(self, field_dict):
        _result = {}
        for key in field_dict.keys():
            _result.update(self.parser_row(key, field_dict[key]))
        return _result

    def parser_row(self, row_flag, row):
        _result_ = {}
        if row_flag == "chl_time":
            ret = self.second_of_week(row)
            _result_ = {row_flag: ret}        # float

        elif row_flag == "GGA" or row_flag == "GGAKF" or row_flag == "GFM":
            GGA = row.split(',')
            if GGA[6] == '1':
                alt = float(GGA[9])         # 海拔
                alt_GS = float(GGA[11])     # 高程异常
                high = alt + alt_GS         # 椭球高
                lat = LogParser.GGA_ll_to_float(GGA[2])
                lon = LogParser.GGA_ll_to_float(GGA[4])

                xyz = lla_to_xyz(lat, lon, high)
                _result_ = {row_flag: {"lat": lat, "lon": lon, "high": high, "xyz": xyz, 'valid': 1}}
            else:
                _result_ = {row_flag: {'valid': 0}}   # {}

        elif row_flag == "RMC" or row_flag == "RMCKF":
            # speed = float(re.findall(r"\d+\.?\d*", row)[3]) * 0.514
            RMC = row.split(',')
            if RMC[7]:
                speed = float(RMC[7]) * 0.514   # float
            else:
                speed = None
            if RMC[2] == 'A':
                _result_ = {row_flag: {"speed": speed, 'valid': 1}}
            else:
                _result_ = {row_flag: {"speed": speed, 'valid': 0}}   # {}

        elif row_flag == "cnr" or row_flag == "pli" or row_flag == "svINFO" or row_flag == "prnNOW":
            ret = re.findall(r"\d+", row)       # [char, ]
            ret_arr = np.array([int(i) for i in ret])
            _result_ = {row_flag: ret_arr}

        elif row_flag == "dopp" or row_flag == "PR":
            ret = re.findall(r"-?\d+\.?\d*", row[:row.index("ir")])
            _result_ = {row_flag: [float(i) for i in ret]}    # [float, ]

        elif row_flag == "fix_sv":
            ret = row[row.index(self.row_flag_dict[row_flag]): row.index(',')]
            _result_ = {row_flag: int(ret)}       # int

        return _result_

    def second_of_week(self, string):
        """获取本周内的秒计数, 如果跨周, 就加上60480"""
        _time_sec = LogParser.chl_time_parse(string)
        if _time_sec < 100 or self.time_bak - _time_sec > 302400:
            if self.time_bak != -1:
                if self.restart:  # 重启
                    if self.restart == 1:  # 重启 后 又 set time
                        if np.fabs(_time_sec - self.time_bak) < 100:
                            self.restart = 0
                    return -1
                _time_sec += 604800  # 没有重启 说明到下一周了
            else:
                return -1  # 板子刚上电

        if self.restart == 1:  # 重启 后 又 set time
            if np.fabs(_time_sec - self.time_bak) < 100:
                self.restart = 0
            return -1
        self.time_bak = _time_sec
        return _time_sec

    def _transpose_(self):
        for key in self.target_row:
            self.all_info_dict[key] = []
        for info in self.all_info_list:
            for key in info.keys():
                self.all_info_dict[key].append(info[key])
            if "prnNOW" not in info.keys():
                self.all_info_dict["prnNOW"].append([])
            if "svINFO" not in info.keys():
                self.all_info_dict["svINFO"].append([])

    @staticmethod
    def chl_time_parse(row):
        _ret = row.split(',')
        return round(int(_ret[1]) / 1000)

    @staticmethod
    def GGA_ll_to_float(ll_str):
        """
        :param ll_str:  GGA经纬度字符串
        :return: 转换GGA经纬度字符串 为 浮点数 (unit 度
        """
        ll_int_part = float(ll_str[: ll_str.find('.') - 2])
        ll_min_part = float(ll_str[ll_str.find('.') - 2:]) / 60.0
        return ll_int_part + ll_min_part


def delete_file(pathname):
    if os.path.exists(pathname):  # 如果文件存在
        # 删除文件，可使用以下两种方法。
        os.remove(pathname)
        # os.unlink(path)   # 删除一个正在使用的文件会报错
    else:
        print('no such file:%s' % pathname)  # 则返回文件不存在


def chart_init(_path_):
    delete_file(_path_ + 'chart/_compare_dopp.xlsx')
    delete_file(_path_ + 'chart/_compare_PR.xlsx')
    delete_file(_path_ + 'chart/_compare_cnr.xlsx')
    fd_Summary_Table = open(_path_ + 'chart/summary_table.md', 'w')
    print("\n## " + _path_.split('/')[-2], file=fd_Summary_Table)
    print("log||final|||||pos||||||vel|||||pli|| |cnr||||PR|||dopp||", file=fd_Summary_Table)
    print(":-|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|"
          ":-:|:-:|:-:|:-:|:-:|:-:|", file=fd_Summary_Table)
    print(".|Sep|Cep|sep50|sep95|sep99|sep std|cep50|cep95|cep99|cep std|v50|v95|v99|v std|fix rate|warning rate|"
          "mean|std|abnormal rate|mean|std|diff mean|diff std|"
          "mean[mean[diff_PR - diff_mean_PR]]|std[mean[diff_PR - diff_mean_PR]]|abnormal rate (100)|"
          "mean[mean[diff_dopp - diff_mean_dopp]]|std[mean[diff_dopp - diff_mean_dopp]]|abnormal rate (5)",
          file=fd_Summary_Table)
    return fd_Summary_Table


if __name__ == '__main__':
    # path = "/home/kwq/work/east_window/0302_night/"
    # file_lst = ["1_mdl5daa_memDbg_east.log"]
    path = "/home/kwq/work/out_test/0219/tt/"
    # file_lst = ["1_mdl_5daa_newFrm_park.log", "COM7_210219_074646_F9P.txt"]
    file_lst = [f for f in os.listdir(path) if f.endswith('.log') or f.endswith('DAT')]
    purpose = {"cnr": ["mean", "std"], "pli": ["mean"], "pos": ["cep50", "cep95", "cep99", "mean", "std"],
               "PR": ["cmp"], "dopp": ["cmp"]}

    ubx_txt = "/home/kwq/work/out_test/0219/tt/COM7_210219_083116_F9P.txt"
    ubx_gga = "/home/kwq/work/out_test/0219/tt/nmea/COM7_210219_083116_F9P.gga"
    # delete_file(path + 'chart/_compare_dopp.xlsx')
    # delete_file(path + 'chart/_compare_PR.xlsx')
    # delete_file(path + 'chart/_compare_cnr.xlsx')
    # fd_summary_table = open(path + 'chart/summary_table.md', 'w')
    # print("log|final_sep|final_cep|||||pos||||||vel|||||pli|| |cnr||||PR|||dopp||", file=fd_summary_table)
    # print(":-|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|:-:|"
    #       ":-:|:-:|:-:|:-:|:-:|:-:|", file=fd_summary_table)
    # print(".|sep50|sep95|sep99|sep std|cep50|cep95|cep99|cep std|v50|v95|v99|v std|fix rate|warning rate|"
    #       "mean|std|abnormal rate|mean|std|diff mean|diff std|"
    #       "mean[mean[diff_PR - diff_mean_PR]]|std[mean[diff_PR - diff_mean_PR]]|abnormal rate (100)|"
    #       "mean[mean[diff_dopp - diff_mean_dopp]]|std[mean[diff_dopp - diff_mean_dopp]]|abnormal rate (5)",
    #       file=fd_summary_table)
    fd_summary_table = chart_init(path)
    Txyz, Tlla, mean_err_dis, std_err_dis = calc_True_Txyz(ubx_gga)
    print("ubx position =", Txyz, Tlla)
    for file in file_lst:
        print(file, end='|', file=fd_summary_table)
        test = LogParser(path+file, purpose, ubx_txt, fd_summary_table)

        start_time = time.time()  # 开始时间
        test.parser_file()
        end_time = time.time()  # 结束时间
        print("耗时: %d" % (end_time - start_time))

        '''进行的操作操作'''
        test.final_pos_analysis(Txyz)
        test.static_pos_cmp(Txyz)
        test.pli_abnormal_pli_mean_cnr_mean()
        test.cnr_cmp()
        test.pr_cmp()
        test.dopp_cmp()
        print("", file=fd_summary_table)
